import sys
import requests
import getopt
import json
import urllib.parse
from veracode_api_signing.plugin_requests import RequestsAuthPluginVeracodeHMAC
import openpyxl
import time
import xml.etree.ElementTree as ET  # for parsing XML
import copy

from veracode_api_signing.credentials import get_credentials

class NoExactMatchFoundException(Exception):
    message=""
    def __init__(self, message_to_set):
        self.message = message_to_set

    def get_message(self):
        return self.message

class Application:
    application_name = ""
    business_criticality = ""
    policy = ""
    description = ""
    tags = ""
    business_unit = ""
    business_owner = ""
    business_owner_email = ""
    teams = []
    dynamic_scan_approval = ""
    archer_application_name = ""
    custom_fields = {}
    collection_name = ""

    def __init__(self, application_name, business_criticality, policy, description, tags, business_unit, business_owner, business_owner_email, teams, dynamic_scan_approval, archer_application_name, custom_fields, collection_name):
        self.application_name = application_name
        self.business_criticality = business_criticality
        self.policy = policy
        self.description = description
        self.tags = tags
        self.business_unit = business_unit
        self.business_owner = business_owner
        self.business_owner_email = business_owner_email
        self.teams = teams
        self.dynamic_scan_approval = dynamic_scan_approval
        self.archer_application_name = archer_application_name
        self.custom_fields = custom_fields
        self.collection_name = collection_name

    def get_archer_application_name_json(self):
        return f''',
            "archer_app_name": "{self.archer_application_name}"'''
    
    def get_business_owner_json(self):
        if not self.business_owner or not self.business_owner_email:
            return f''',
        "business_owners": []'''
        return f''',
        "business_owners": [
            {{
                "email": "{self.business_owner_email}",
                "name": "{self.business_owner}"
            }}
        ]'''
    
    def get_business_unit_json(self):
        if not self.business_unit:
            return f''',
            "business_unit": null'''
        return f''',
        "business_unit": {{
            "guid": "{self.business_unit}"
        }}'''
    
    def get_policy_json(self):
        if not self.policy:
            return f''',
            "policies": []'''
        return f''',
        "policies": [{{
            "guid": "{self.policy}"
        }}]'''
    

    def get_teams_json(self):
        inner_team_list = ""
        for team_guid in self.teams:
            inner_team_list = inner_team_list + (""",
            """ if inner_team_list else "") + f'{{ "guid": "{team_guid}" }}'
        if inner_team_list:
            return f''',
                "teams": [
                    {inner_team_list}
                ]'''
        else:
            return ""
    
    def get_application_settings_json(self):
        value = False
        if self.dynamic_scan_approval:
            value = str(self.dynamic_scan_approval).strip().lower() == "false"
        return f''',
        "settings": {{
            "dynamic_scan_approval_not_required": {str(value).lower()}
        }}'''
    
    def get_custom_fields_json(self):
        inner_custom_fields_list = ""
        for name, value in self.custom_fields.items():
            new_field = f'''{{
                "name": "{name}",
                "value": "{value}"
            }}'''

            inner_custom_fields_list = inner_custom_fields_list + (""",
                """ if inner_custom_fields_list else "") + new_field
        if inner_custom_fields_list:
            return f''',
                    "custom_fields": [
                        {inner_custom_fields_list}
                    ]'''
        else:
            return ""

class Created_application:
    guid = ""

    def __init__(self, guid):
        self.guid = guid

class Created_collection:
    guid = ""

    def __init__(self, guid):
        self.guid = guid

class Failure_to_create:
    error = ""

    def __init__(self, error):
        self.error = error

json_headers = {
    "User-Agent": "Bulk application creation - python script",
    "Content-Type": "application/json"
}

failed_attempts = 0
max_attempts_per_request = 10
sleep_time = 10

last_column = 0
non_custom_field_headers={"Application Name",
                            "Business Criticality",
                            "Policy",
                            "Description",
                            "Tags",
                            "Business Unit",
                            "Business Owner",
                            "Owner Email",
                            "Teams",
                            "Dynamic Scan Approval",
                            "Archer Application Name",
                            "Collection Name"}

DEFAULT_SEPARATOR = ": "
DEFAULT_HEADER_ROW = 2
NEW_APPLICATION_NAME_ROW_PREFIX = "newApp"
TOTAL_FAILURE = "Total failure"
NULL = "NULL"

def print_help():
    """Prints command line options and exits"""
    print("""veracode-application-profile-splitting-helper.py -f <excel_file_with_application_definitions> [-s <separator>] [-r <header_row>] [-d]"
        Reads all lines in <excel_file_with_application_definitions>, for each line, it will update the profile
        <header_row> defines which row contains your table headers, which will be read to determine where each field goes (default 2).
        <separator> defines the separator to be used to name the new application profiles (default ': ').
""")
    sys.exit()

def request_encode(value_to_encode):
    return urllib.parse.quote(value_to_encode, safe='')

def find_exact_match(list, to_find, field_name, list_name2):
    if list_name2:
        for index in range(len(list)):
            if (list_name2 and list[index][list_name2][field_name].lower() == to_find.lower()):
                return list[index]
    
        print(f"Unable to find a member of list with {field_name}+{list_name2} equal to {to_find}")
        raise NoExactMatchFoundException(f"Unable to find a member of list with {field_name}+{list_name2} equal to {to_find}")
    else:
        for index in range(len(list)):
            if list[index][field_name].lower() == to_find.lower():
                return list[index]

        print(f"Unable to find a member of list with {field_name} equal to {to_find}")
        raise NoExactMatchFoundException(f"Unable to find a member of list with {field_name} equal to {to_find}")

def get_field_value(excel_headers, excel_sheet, row, field_header):
    field_to_get = field_header.strip()
    if field_to_get in excel_headers:
        field_value = excel_sheet.cell(row = row, column = excel_headers[field_to_get]).value
        if field_value:
            return field_value
    return ""

def get_business_owners(excel_headers, excel_sheet, row):
    name=get_field_value(excel_headers, excel_sheet, row, "Business Owner")
    email=get_field_value(excel_headers, excel_sheet, row, "Owner Email")
    if not name or not email:
        return ""
    return f''',
    "business_owners": [
      {{
        "email": "{email}",
        "name": "{name}"
      }}
    ]'''

def get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, list_name2, field_to_check, field_to_get, is_exact_match, verbose):
    global failed_attempts
    global sleep_time
    global max_attempts_per_request
    path = f"{api_base}{api_to_call}"
    if verbose:
        print(f"Calling: {path}")

    response = requests.get(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers)
    data = response.json()

    if response.status_code == 200:
        if verbose:
            print(data)
        if "_embedded" in data and len(data["_embedded"][list_name]) > 0:
            found_match = None
            if list_name2:
                found_match = (find_exact_match(data["_embedded"][list_name], item_to_find, field_to_check, list_name2) if is_exact_match else data["_embedded"][list_name][list_name2][0])
            else:
                found_match = (find_exact_match(data["_embedded"][list_name], item_to_find, field_to_check, list_name2) if is_exact_match else data["_embedded"][list_name][0])
            return found_match[field_to_get] if field_to_get else found_match
        else:
            print(f"ERROR: No {list_name}+{list_name2} named '{item_to_find}' found")
            return f"ERROR: No {list_name}+{list_name2} named '{item_to_find}' found"
    else:
        print(f"ERROR: trying to get {list_name}+{list_name2} named {item_to_find}")
        print(f"ERROR: code: {response.status_code}")
        print(f"ERROR: value: {data}")
        failed_attempts+=1
        if (failed_attempts < max_attempts_per_request):
            time.sleep(sleep_time)
            return get_item_from_api_call(api_base, api_to_call, item_to_find, list_name, list_name2, field_to_check, field_to_get, verbose)
        else:
            return f"ERROR: trying to get {list_name}+{list_name2} named {item_to_find}"

def get_business_unit(api_base, business_unit_name, verbose):
    if not business_unit_name:
        return ""
    else:
        return get_item_from_api_call(api_base, "api/authn/v2/business_units?bu_name="+ request_encode(business_unit_name), business_unit_name, "business_units", None, "bu_name", "bu_id", True, verbose)

def get_policy(api_base, policy_name, verbose):
    return get_item_from_api_call(api_base, "appsec/v1/policies?category=APPLICATION&name_exact=true&public_policy=true&name="+ request_encode(policy_name), policy_name, "policy_versions", None, "name", "guid", False, verbose)

def get_team_guid(api_base, team_name, verbose):
    return get_item_from_api_call(api_base, "api/authn/v2/teams?all_for_org=true&team_name="+ request_encode(team_name), team_name, "teams", None, "team_name", "team_id", True, verbose)

def get_teams(api_base, teams_base, verbose):
    all_teams = teams_base.split(",")
    team_list = []
    for team_name in all_teams:
        team_guid = get_team_guid(api_base, team_name.strip(), verbose)
        if team_guid: 
            team_list.append(team_guid)
    return team_list
        
def get_application_settings(excel_headers, excel_sheet, row):
    base_value = get_field_value(excel_headers, excel_sheet, row, "Dynamic Scan Approval")
    value = False
    if base_value:
        value = str(base_value).strip().lower() == "false"
    return f''',
    "settings": {{
      "dynamic_scan_approval_not_required": {str(value).lower()}
    }}'''

def get_custom_fields(excel_headers, excel_sheet, row):
    global non_custom_field_headers
    custom_field_dict = {}
    for field in excel_headers:
        if not field.startswith(NEW_APPLICATION_NAME_ROW_PREFIX) and not field in non_custom_field_headers:
            value = excel_sheet.cell(row = row, column=excel_headers[field]).value
            if value:
                custom_field_dict[field] = value
    return dict(custom_field_dict)
        
def get_archer_application_name(excel_headers, excel_sheet, row):
    archer_app_name = get_field_value(excel_headers, excel_sheet, row, "Archer Application Name")
    if archer_app_name:
        return f''',
        "archer_app_name": "{archer_app_name}"'''
    else:
        return ""

def url_encode_with_plus(a_string):
    return urllib.parse.quote_plus(a_string, safe='').replace("&", "%26")

def get_error_node_value(body):
    inner_node = ET.XML(body)
    if inner_node.tag == "error" and not inner_node == None:
        return inner_node.text
    else:
        return ""
    

def create_application(api_base, applications_to_create: Application, verbose):
    path = f"{api_base}appsec/v1/applications"
    request_content=f'''{{
        "profile": {{
            "business_criticality": "{applications_to_create.business_criticality}"            
            {applications_to_create.get_archer_application_name_json()}
            {applications_to_create.get_business_owner_json()}
            {applications_to_create.get_business_unit_json()},
            "description": "{applications_to_create.description}",
            "name": "{applications_to_create.application_name}"
            {applications_to_create.get_policy_json()},
            "tags": "{applications_to_create.tags}"
            {applications_to_create.get_teams_json()}
            {applications_to_create.get_application_settings_json()}
            {applications_to_create.get_custom_fields_json()}
        }}
    }}'''
    if verbose:
        print(request_content)

    response = requests.post(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content))

    body = response.json()
    if verbose:
        print(f"status code {response.status_code}")
        if body:
            print(body)
    if response.status_code == 200:
        print(f"Successfully created application profile '{applications_to_create.application_name}.")
        return Created_application(body["guid"])
    else:
        body = response.json()
        if (body):
            return Failure_to_create(f"Unable to create application profile: {response.status_code} - {body}")
        else:
            return Failure_to_create(f"Unable to create application profile: {response.status_code}")
    

def setup_excel_headers(excel_sheet, header_row, verbose):
    excel_headers = {}
    global last_column
    for column in range(1, excel_sheet.max_column+1):
        cell = excel_sheet.cell(row = header_row, column = column)
        if not cell or cell is None or not cell.value or cell.value.strip() == "":
            break
        to_add = cell.value
        if to_add:
            to_add = str(to_add).strip()
        if verbose:
            print(f"Adding column {column} for value {to_add}")
        excel_headers[to_add] = column
        last_column += 1
    return excel_headers

def build_custom_fields_from_base_json(custom_fields_json):
    if not custom_fields_json:
        return {}
    
    names = []
    values = []
    for custom_field in custom_fields_json:
        names.append(custom_field["name"])
        values.append(custom_field["value"])
        
    return dict(zip(names, values))

def get_base_application(api_base, application_name, verbose):
    base_application_json = get_item_from_api_call(api_base, "appsec/v1/applications?name="+ request_encode(application_name.strip()), application_name.strip(), "applications", "profile", "name", None, True, verbose)
    if not base_application_json:
        return None
    business_owner = base_application_json["profile"]["business_owners"]
    business_owner = business_owner[0] if business_owner else None
    return Application(application_name = application_name,
                        business_criticality = base_application_json["profile"]["business_criticality"],
                        policy = base_application_json["profile"]["policies"][0]["guid"],
                        description = base_application_json["profile"]["description"],
                        tags = base_application_json["profile"]["tags"],
                        business_unit = base_application_json["profile"]["business_unit"]["guid"],
                        business_owner = business_owner["name"] if business_owner else "",
                        business_owner_email = business_owner["email"] if business_owner else "",
                        teams = map(lambda team: team["guid"], base_application_json["profile"]["teams"]),
                        dynamic_scan_approval = base_application_json["profile"]["settings"]["dynamic_scan_approval_not_required"],
                        archer_application_name = base_application_json["profile"]["archer_app_name"],
                        custom_fields = build_custom_fields_from_base_json(base_application_json["profile"]["custom_fields"]),
                        collection_name = "")

def get_all_applications_to_create(base_application, separator, excel_headers, excel_sheet, row):
    applications_to_create = []
    for field in excel_headers:
        if field.startswith(NEW_APPLICATION_NAME_ROW_PREFIX):
            value = excel_sheet.cell(row = row, column=excel_headers[field]).value
            if value:
                new_application = copy.deepcopy(base_application)
                new_application.application_name = base_application.application_name + separator + value
                applications_to_create.append(new_application)
    return applications_to_create

def get_value_from_excel_or_base(base_value, excel_value):
    if excel_value:
        return "" if excel_value == NULL else excel_value
    return base_value

def get_policy_from_excel_or_base(api_base, base_value, policy_name, verbose):
    if policy_name:
        return "" if policy_name == NULL else get_policy(api_base, policy_name, verbose)
    return base_value

def get_business_unit_from_excel_or_base(api_base, base_value, business_unit_name, verbose):
    if business_unit_name:
        return "" if business_unit_name == NULL else get_business_unit(api_base, business_unit_name, verbose)
    return base_value

def get_teams_from_excel_or_base(api_base, base_value, teams_base, verbose):
    if teams_base:
        return "" if teams_base == NULL else get_teams(api_base, teams_base, verbose)
    return base_value

def combine_custom_fields(base_custom_fields, excel_headers, excel_sheet, row):
    new_custom_fields = get_custom_fields(excel_headers, excel_sheet, row)
    for field, value in new_custom_fields.items():
        is_null = value == NULL
        if is_null:
            base_custom_fields[field] = ""
        else:
            base_custom_fields[field] = value

    return base_custom_fields

def set_values_from_excel(api_base, base_application, excel_headers, excel_sheet, row, verbose):
    base_application.business_criticality = get_value_from_excel_or_base(base_application.business_criticality, get_field_value(excel_headers, excel_sheet, row, "Business Criticality")).upper()
    
    base_application.description = get_value_from_excel_or_base(base_application.description, get_field_value(excel_headers, excel_sheet, row, "Description"))
    base_application.tags = get_value_from_excel_or_base(base_application.tags, get_field_value(excel_headers, excel_sheet, row, "Tags"))
    base_application.business_owner = get_value_from_excel_or_base(base_application.business_owner, get_field_value(excel_headers, excel_sheet, row, "Business Owner"))
    base_application.business_owner_email = get_value_from_excel_or_base(base_application.business_owner_email, get_field_value(excel_headers, excel_sheet, row, "Owner Email"))
    base_application.dynamic_scan_approval = get_value_from_excel_or_base(base_application.dynamic_scan_approval, get_field_value(excel_headers, excel_sheet, row, "Dynamic Scan Approval"))
    base_application.archer_application_name = get_value_from_excel_or_base(base_application.archer_application_name, get_field_value(excel_headers, excel_sheet, row, "Archer Application Name"))

    base_application.custom_fields = combine_custom_fields(base_application.custom_fields, excel_headers, excel_sheet, row)
    base_application.policy = get_policy_from_excel_or_base(api_base, base_application.policy, get_field_value(excel_headers, excel_sheet, row, "Policy"), verbose)
    base_application.business_unit = get_business_unit_from_excel_or_base(api_base, base_application.business_unit, get_field_value(excel_headers, excel_sheet, row, "Business Unit"), verbose)
    base_application.teams = get_teams_from_excel_or_base(api_base, base_application.teams, get_field_value(excel_headers, excel_sheet, row, "Teams"), verbose)
    
    return base_application

def parse_created_assets(created_guid_array):
    result = ""
    for new_guid in created_guid_array:
        if result:
            result = result + ''',
            '''
        result = result + f'''{{
                "guid": "{new_guid}",
                "type": "APPLICATION"
            }}'''
    return result

def create_collection(api_base, created_guid_array, original_base_application, verbose):
    path = f"{api_base}appsec/v1/collections/"
    request_content=f'''{{
        "name": "{original_base_application.collection_name}",
        "description": "{original_base_application.description}",
        "tags": "{original_base_application.tags}",
        "business_unit": {{
            "guid": "{original_base_application.business_unit}"
        }}
        {original_base_application.get_custom_fields_json()},
        "asset_infos": [{parse_created_assets(created_guid_array)}]
    }}'''
    if verbose:
        print(request_content)

    response = requests.post(path, auth=RequestsAuthPluginVeracodeHMAC(), headers=json_headers, json=json.loads(request_content))


    body = response.json()
    if verbose:
        print(f"status code {response.status_code}")
        if body:
            print(body)
    if response.status_code == 200:
        print(f"Successfully created collection '{original_base_application.collection_name}'.")
        return Created_collection(body["guid"])
    else:
        body = response.json()
        if (body):
            return Failure_to_create(f"Unable to create collection: {response.status_code} - {body}")
        else:
            return Failure_to_create(f"Unable to create collection: {response.status_code}")

def split_application(api_base, separator, excel_headers, excel_sheet, row, verbose):
    base_application_name = get_field_value(excel_headers, excel_sheet, row, "Application Name")
    if not base_application_name:
        print("Skipping empty line")
        return "", ""
    print(f"Trying to split application '{base_application_name}'")
    base_application = get_base_application(api_base, base_application_name, verbose)
    if not base_application:
        error_message = f"Unable to find application named {base_application_name}"
        print(error_message)
        return (TOTAL_FAILURE, error_message)

    base_application.collection_name = get_field_value(excel_headers, excel_sheet, row, "Collection Name")
    original_base_application = copy.deepcopy(base_application)

    base_application = set_values_from_excel(api_base, base_application, excel_headers, excel_sheet, row, verbose)

    applications_to_create = get_all_applications_to_create(base_application, separator, excel_headers, excel_sheet, row)

    should_create_collection: bool = base_application.collection_name

    errors_array = []
    created_guid_array = []
    has_at_least_one_success = False
    for application in applications_to_create:
        result = create_application(api_base, application, verbose)
        if isinstance(result, Failure_to_create):
            errors_array.append(result.error)
        else:
            created_guid_array.append(result.guid)
            has_at_least_one_success = True
    
    
    if should_create_collection and has_at_least_one_success:
        result = create_collection(api_base, created_guid_array, original_base_application, verbose)
        if isinstance(result, Failure_to_create):
            errors_array.append(result.error)
    
    errors = " - ERROR: ".join(errors_array) if errors_array else ""

    print("-----------------------------------")
    if has_at_least_one_success:
        return ("success", errors)
    else:
        return (TOTAL_FAILURE, errors)


def create_all_applications(api_base, file_name, header_row, separator, verbose):
    global failed_attempts
    excel_file = openpyxl.load_workbook(file_name)
    excel_sheet = excel_file.active
    try:
        excel_headers = setup_excel_headers(excel_sheet, header_row, verbose)
        print("Finished reading excel headers")
        if verbose:
            print("Values found are:")
            print(excel_headers)

        max_column=len(excel_headers)
        for row in range(header_row+1, excel_sheet.max_row+1):      
            failed_attempts = 0
            if verbose:
                for field in excel_headers:
                    print(f"Found column with values:")
                    print(f"{field} -> {excel_sheet.cell(row = row, column=excel_headers[field]).value}")
            status=excel_sheet.cell(row = row, column = max_column+1).value
            errors = ""
            if (status == 'success'):
                print("Skipping row as it was already done")
            else:
                try:
                    print(f"Importing row {row-header_row}/{excel_sheet.max_row-header_row}:")
                    status, errors = split_application(api_base, separator, excel_headers, excel_sheet, row, verbose)
                    print(f"Finished importing row {row-header_row}/{excel_sheet.max_row-header_row}")
                    print("---------------------------------------------------------------------------")
                except NoExactMatchFoundException:
                    status = TOTAL_FAILURE
                    errors = NoExactMatchFoundException.get_message()
                excel_sheet.cell(row = row, column = max_column+1).value=status
                excel_sheet.cell(row = row, column = max_column+2).value=errors
    finally:
        excel_file.save(filename=file_name)

def get_api_base():
    api_key_id, api_key_secret = get_credentials()
    api_base = "https://api.veracode.{instance}/"
    if api_key_id.startswith("vera01"):
        return api_base.replace("{instance}", "eu", 1)
    else:
        return api_base.replace("{instance}", "com", 1)

def main(argv):
    """Allows for bulk adding application profiles"""
    global failed_attempts
    global last_column
    excel_file = None
    try:
        verbose = False
        file_name = ''
        header_row = -1
        separator = ''

        opts, args = getopt.getopt(argv, "hdf:r:s:", ["file_name=", "header_row=", "separator="])
        for opt, arg in opts:
            if opt == '-h':
                print_help()
            if opt == '-d':
                verbose = True
            if opt in ('-f', '--file_name'):
                file_name=arg
            if opt in ('-s', '--separator'):
                separator=arg
            if opt in ('-r', '--header_row'):
                header_row=int(arg)

        api_base = get_api_base()        
        if file_name:
            if not separator:
                print(f"Using default separator {DEFAULT_SEPARATOR}")
                separator = DEFAULT_SEPARATOR
            if header_row < 1:
                print(f"Using default header row {DEFAULT_HEADER_ROW}")
                header_row = DEFAULT_HEADER_ROW
            create_all_applications(api_base, file_name, header_row, separator, verbose)
        else:
            print_help()
    except requests.RequestException as e:
        print("An error occurred!")
        print(e)
        sys.exit(1)
    finally:
        if excel_file:
            excel_file.save(filename=file_name)


if __name__ == "__main__":
    main(sys.argv[1:])
